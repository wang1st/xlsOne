#!/usr/bin/env python3
"""
xlsOne CLI - Excel 报表合并命令行工具 v0.3.0
支持多 sheet 解析和同名 sheet 汇总
"""

import os
import sys
import json
import glob
import argparse
from pathlib import Path
from typing import Optional, Dict, List
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent / "xlsonecore"))

from xlsonecore import (
    ExcelParser, MergerEngine, MergeStrategy,
    NumericMergeStrategy, TextMergeStrategy, CodeMergeStrategy
)


def print_header():
    print("🔷 xlsOne - Excel 报表合并工具 v0.3.0")
    print("=" * 50)


def print_stats(stats):
    print(f"\n📈 统计: {stats.total_files} 文件, {stats.total_cells} 单元格")


def print_preview(merged_cells, max_rows: int = 5):
    print(f"\n📋 预览 (前{max_rows}行):")
    print("-" * 60)
    for i, row in enumerate(merged_cells[:max_rows]):
        row_output = "  "
        for cell in row:
            display_value = cell.result_value if cell.result_value else "(空)"
            row_output += f"{display_value:<15}  "
        print(row_output)
    print("-" * 60)


def export_all_sheets_to_json(all_merged: Dict[str, tuple], output_path: str):
    """导出所有 sheet 为 JSON"""
    data = {"sheets": {}}
    
    for sheet_name, (merged_cells, stats) in all_merged.items():
        data["sheets"][sheet_name] = {
            "statistics": {
                "totalFiles": stats.total_files,
                "totalCells": stats.total_cells,
                "numericCells": stats.numeric_cells,
                "textCells": stats.text_cells,
                "codeCells": stats.code_cells,
                "emptyCells": stats.empty_cells
            },
            "cells": [
                [
                    {
                        "position": {
                            "row": cell.position.row,
                            "column": cell.position.column,
                            "a1Notation": cell.position.a1_notation()
                        },
                        "resultType": cell.result_type,
                        "resultValue": cell.result_value,
                        "numericValue": cell.numeric_value,
                        "sourceValues": cell.source_values
                    }
                    for cell in row
                ]
                for row in merged_cells
            ]
        }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 导出 JSON: {output_path}")


def export_to_xlsx(all_merged: Dict[str, tuple], output_path: str):
    """导出所有 sheet 为 Excel"""
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # 移除默认 sheet
    default_sheet = wb.active
    
    for sheet_name, (merged_cells, stats) in all_merged.items():
        # 创建新 sheet
        if sheet_name == list(all_merged.keys())[0]:
            ws = wb.active
            ws.title = sheet_name[:31]  # Excel sheet 名最长31字符
        else:
            ws = wb.create_sheet(title=sheet_name[:31])
        
        # 写入数据
        for row_idx, row in enumerate(merged_cells, start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.result_value is not None:
                    # 尝试写入数值
                    try:
                        ws.cell(row=row_idx, column=col_idx, value=float(cell.result_value))
                    except ValueError:
                        ws.cell(row=row_idx, column=col_idx, value=cell.result_value)
    
    wb.save(output_path)
    print(f"\n💾 导出 Excel: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="xlsOne - Excel 多 Sheet 报表合并工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s *.xlsx                           # 合并所有文件的同名 sheet
  %(prog)s "仙居县/*.xlsx" --all-sheets     # 解析所有 sheet 并按名汇总
  %(prog)s *.xlsx -o result.json           # 导出为 JSON
        """
    )
    
    parser.add_argument('files', nargs='*', help='Excel 文件路径')
    parser.add_argument('-o', '--output', help='输出文件路径')
    parser.add_argument('-f', '--format', choices=['json', 'xlsx'], default='xlsx', help='输出格式 (默认: xlsx)')
    parser.add_argument('-n', '--numeric-strategy', choices=['sum', 'average', 'max', 'min', 'first', 'last'], default='sum')
    parser.add_argument('-t', '--text-strategy', choices=['majority', 'first', 'last', 'concat'], default='majority')
    parser.add_argument('-c', '--code-strategy', choices=['placeholder', 'first', 'last'], default='placeholder')
    parser.add_argument('--all-sheets', action='store_true', help='解析所有 sheet 并按名汇总')
    parser.add_argument('--stats-only', action='store_true', help='只显示统计')
    parser.add_argument('-v', '--verbose', action='store_true', help='详细输出')
    
    args = parser.parse_args()
    
    print_header()
    
    # 收集文件
    file_paths: list[str] = []
    for pattern in args.files:
        if '*' in pattern or '?' in pattern:
            file_paths.extend(glob.glob(pattern))
        elif os.path.exists(pattern):
            file_paths.append(pattern)
    
    file_paths = list(dict.fromkeys(file_paths))
    
    if not file_paths:
        print("❌ 错误: 没有找到文件")
        return 1
    
    print(f"📁 {len(file_paths)} 个文件")
    
    # 初始化
    parser = ExcelParser()
    strategy = MergeStrategy(
        numeric_strategy=NumericMergeStrategy(args.numeric_strategy),
        text_strategy=TextMergeStrategy(args.text_strategy),
        code_strategy=CodeMergeStrategy(args.code_strategy)
    )
    engine = MergerEngine(strategy=strategy)
    
    if args.all_sheets:
        # === 解析所有 sheet 并按名称汇总 ===
        print("\n📊 解析所有 sheet...")
        
        # 按 sheet 名称分组数据
        sheets_data: Dict[str, List] = defaultdict(list)
        
        for file_path in file_paths:
            try:
                all_sheets = parser.parse_all_sheets(file_path)
                for sheet_name, data in all_sheets.items():
                    sheets_data[sheet_name].append(data)
                if args.verbose:
                    print(f"   ✓ {os.path.basename(file_path)}: {len(all_sheets)} sheets")
            except Exception as e:
                print(f"❌ 失败: {os.path.basename(file_path)} - {e}")
        
        # 汇总每个 sheet
        print("\n🔄 汇总同名 sheets...")
        all_merged: Dict[str, tuple] = {}
        
        for sheet_name, data_list in sheets_data.items():
            try:
                merged_cells, stats = engine.merge(data_list)
                all_merged[sheet_name] = (merged_cells, stats)
                print(f"   ✓ {sheet_name}: {stats.total_cells} 单元格")
            except ValueError as e:
                print(f"   ⚠️ {sheet_name}: 跳过 - {e}")
        
        # 输出结果
        print(f"\n✅ 完成! 汇总了 {len(all_merged)} 个 sheets:")
        for sheet_name, (_, stats) in all_merged.items():
            print(f"   - {sheet_name}: {stats.total_cells} 单元格, {stats.numeric_cells} 数值")
        
        # 导出
        output_path = args.output or f"merged_result.{args.format}"
        
        if args.format == 'xlsx':
            export_to_xlsx(all_merged, output_path)
        else:
            export_all_sheets_to_json(all_merged, output_path)
        
        # 预览第一个 sheet
        if not args.stats_only and all_merged:
            first_sheet = list(all_merged.keys())[0]
            print_preview(all_merged[first_sheet][0])
    
    else:
        # === 只解析第一个 sheet ===
        print("\n📊 解析第一个 sheet...")
        
        all_data: List[List] = []
        for file_path in file_paths:
            try:
                data = parser.parse(file_path)
                all_data.append(data)
                if args.verbose:
                    print(f"   ✓ {os.path.basename(file_path)}: {len(data)} 行")
            except Exception as e:
                print(f"❌ 失败: {os.path.basename(file_path)}")
        
        print("\n🔄 汇总...")
        merged_cells, stats = engine.merge(all_data)
        
        print_stats(stats)
        
        if args.output:
            if args.format == 'xlsx':
                export_to_xlsx({"汇总": (merged_cells, stats)}, args.output)
            else:
                export_all_sheets_to_json({"汇总": (merged_cells, stats)}, args.output)
        elif not args.stats_only:
            output_path = "merged_result.xlsx"
            export_to_xlsx({"汇总": (merged_cells, stats)}, output_path)
            print_preview(merged_cells)
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
