"""
xlsOneCore - Excel 解析器
"""

import os
from typing import Optional
from openpyxl import load_workbook
from . import (
    CellData, CellPosition, CellDataType,
    recognize_data_type, extract_numeric_value
)


class ExcelParser:
    """Excel 文件解析器"""
    
    def __init__(self):
        pass
    
    def parse(self, file_path: str, sheet_name: Optional[str] = None) -> list[list[CellData]]:
        """
        解析单个 Excel 文件
        
        Args:
            file_path: Excel 文件路径
            sheet_name: 工作表名称，默认第一个
            
        Returns:
            二维数组，每个元素是 CellData
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        wb = load_workbook(filename=file_path, data_only=True)
        
        # 选择工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        result: list[list[CellData]] = []
        
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            row_data: list[CellData] = []
            
            for col_idx, cell in enumerate(row, start=1):
                position = CellPosition(row=row_idx - 1, column=col_idx - 1)
                raw_value = self._get_cell_value(cell)
                cell_type = recognize_data_type(raw_value)
                numeric_value = extract_numeric_value(raw_value, cell_type)
                
                cell_data = CellData(
                    position=position,
                    raw_value=raw_value,
                    cell_type=cell_type,
                    numeric_value=numeric_value
                )
                row_data.append(cell_data)
            
            if row_data:
                result.append(row_data)
        
        return result
    
    def parse_all_sheets(self, file_path: str) -> dict[str, list[list[CellData]]]:
        """
        解析所有工作表
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            字典，key 是 sheet 名称，value 是数据
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        wb = load_workbook(filename=file_path, data_only=True)
        all_data: dict[str, list[list[CellData]]] = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data: list[list[CellData]] = []
            
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                row_data: list[CellData] = []
                
                for col_idx, cell in enumerate(row, start=1):
                    position = CellPosition(row=row_idx - 1, column=col_idx - 1)
                    raw_value = self._get_cell_value(cell)
                    cell_type = recognize_data_type(raw_value)
                    numeric_value = extract_numeric_value(raw_value, cell_type)
                    
                    cell_data = CellData(
                        position=position,
                        raw_value=raw_value,
                        cell_type=cell_type,
                        numeric_value=numeric_value
                    )
                    row_data.append(cell_data)
                
                if row_data:
                    data.append(row_data)
            
            all_data[sheet_name] = data
        
        return all_data
    
    def parse_multiple(self, file_paths: list[str]) -> list[list[list[CellData]]]:
        """
        解析多个 Excel 文件
        
        Args:
            file_paths: 文件路径列表
            
        Returns:
            数据集列表
        """
        all_data: list[list[list[CellData]]] = []
        
        for file_path in file_paths:
            try:
                data = self.parse(file_path)
                all_data.append(data)
            except Exception as e:
                print(f"⚠️  解析失败 {file_path}: {e}")
        
        return all_data
    
    def _get_cell_value(self, cell) -> Optional[str]:
        """获取单元格值"""
        if cell.value is None:
            return None
        
        value = cell.value
        
        # 处理各种类型
        if isinstance(value, (int, float)):
            return str(value)
        elif isinstance(value, str):
            return value.strip() if value else None
        else:
            return str(value)
    
    def get_sheet_names(self, file_path: str) -> list[str]:
        """获取工作表名称列表"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        wb = load_workbook(filename=file_path, data_only=True)
        return wb.sheetnames


if __name__ == "__main__":
    # 测试解析
    parser = ExcelParser()
    
    # 测试文件
    import glob
    test_files = glob.glob("/Users/ethan/xlsOne/仙居县/*.xlsx")
    
    if test_files:
        print(f"📁 测试文件: {test_files[0]}")
        data = parser.parse(test_files[0])
        print(f"✓ 解析成功: {len(data)} 行")
        print(f"  首行数据: {[cell.raw_value for cell in data[0][:5]]}")
