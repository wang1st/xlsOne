from __future__ import annotations

from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from generate_demo_workbooks import units


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "recording_workbooks"
OUTPUT = ROOT / "template_change_demo"


def main() -> None:
    if OUTPUT.exists():
        shutil.rmtree(OUTPUT)
    OUTPUT.mkdir(parents=True)
    demo_units = units()[:24]
    for unit in demo_units:
        source = SOURCE / f"{unit.name}.xlsx"
        target = OUTPUT / source.name
        shutil.copy2(source, target)

    changed = OUTPUT / f"{demo_units[17].name}.xlsx"
    workbook = load_workbook(changed)
    sheet = workbook["费用明细"]
    new_row = sheet.max_row + 1
    sheet.cell(new_row, 1, "模板外新增行")
    sheet.cell(new_row, 2, "临时费用审批备注")
    sheet.cell(new_row, 3, 128.50)
    for cell in sheet[new_row]:
        cell.fill = PatternFill("solid", fgColor="FFF1F0")
        cell.font = Font(color="CF1322", bold=True)
    workbook.save(changed)

    print(f"created {len(list(OUTPUT.glob('*.xlsx')))} files")
    print(f"changed workbook: {changed.name}, sheet: 费用明细, rows: {sheet.max_row}")


if __name__ == "__main__":
    main()
