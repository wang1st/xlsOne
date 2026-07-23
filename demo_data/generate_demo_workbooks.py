from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
import json
import math
import random
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "workbooks"
RECORDING_OUTPUT = ROOT / "recording_workbooks"
PERIOD = "2026年6月"
SEED = 20260721

REGION_CITIES = {
    "华东": [
        "南京", "苏州", "无锡", "常州", "南通", "扬州", "镇江", "泰州", "徐州", "连云港",
        "杭州", "宁波", "温州", "嘉兴", "绍兴", "金华", "台州", "湖州", "合肥", "芜湖",
    ],
    "华南": [
        "广州", "深圳", "佛山", "东莞", "珠海", "惠州", "中山", "江门", "肇庆", "汕头",
        "湛江", "茂名", "清远", "南宁", "柳州", "桂林", "海口", "三亚", "福州", "厦门",
    ],
    "华北": [
        "石家庄", "唐山", "保定", "廊坊", "沧州", "邯郸", "邢台", "秦皇岛", "衡水", "张家口",
        "太原", "大同", "长治", "晋城", "临汾", "呼和浩特", "包头", "鄂尔多斯", "烟台", "潍坊",
    ],
    "华中": [
        "武汉", "宜昌", "襄阳", "荆州", "黄石", "孝感", "长沙", "株洲", "湘潭", "衡阳",
        "岳阳", "常德", "郑州", "洛阳", "开封", "新乡", "南昌", "赣州", "九江", "上饶",
    ],
    "西南": [
        "成都", "绵阳", "德阳", "乐山", "宜宾", "南充", "泸州", "眉山", "资阳", "遂宁",
        "昆明", "曲靖", "玉溪", "昭通", "贵阳", "遵义", "六盘水", "毕节", "拉萨", "昌都",
    ],
    "西北及东北": [
        "西安", "宝鸡", "咸阳", "渭南", "延安", "汉中", "兰州", "天水", "武威", "张掖",
        "银川", "吴忠", "乌鲁木齐", "克拉玛依", "西宁", "沈阳", "大连", "长春", "哈尔滨", "齐齐哈尔",
    ],
}

BLUE = "1677FF"
DEEP_BLUE = "0B2B4B"
LIGHT_BLUE = "EAF3FF"
LIGHT_GRAY = "F5F7FA"
MID_GRAY = "D9E1EA"
TEXT = "172B4D"
WHITE = "FFFFFF"
RED = "D92D20"
AMBER = "B54708"
GREEN = "067647"
THIN = Side(style="thin", color=MID_GRAY)


@dataclass
class Unit:
    index: int
    code: str
    region: str
    city: str
    name: str
    scale: float


def round_money(value: float) -> float:
    return round(max(0.0, value), 2)


def units() -> list[Unit]:
    result: list[Unit] = []
    index = 1
    for region, cities in REGION_CITIES.items():
        for city in cities:
            scale = 0.72 + ((index * 37) % 89) / 100
            result.append(Unit(index, f"BR{index:03d}", region, city,
                               f"{city}分公司", scale))
            index += 1
    assert len(result) == 120
    return result


ANOMALIES = {
    17: ("销售费用异常", "市场活动集中投放，销售费用明显高于常态", "市场部"),
    43: ("逾期回款", "重点客户验收推迟，90天以上应收款增加", "财务部"),
    68: ("项目延期", "两个交付项目因客户接口调整延期", "交付部"),
    91: ("服务压力", "版本升级导致服务工单阶段性增加", "客户成功部"),
    113: ("经营目标未达成", "重点项目签约推迟，本月收入低于预算", "经营管理部"),
}

ANOMALY_CHECK = {
    17: "销售费用",
    43: "应收回款",
    68: "项目交付",
    91: "服务质量",
    113: "经营目标",
}


def style_sheet(ws, widths: list[float], freeze: str = "A5") -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 24
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=DEEP_BLUE)
        cell.font = Font(name="微软雅黑", size=15, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center")
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(name="微软雅黑", size=10, color=TEXT)
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN)
    ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.font = Font(name="微软雅黑", size=10, color=TEXT)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)


def add_header(ws, title: str, subtitle: str, headers: list[str]) -> None:
    ws.append([title] + [""] * (len(headers) - 1))
    ws.append([subtitle] + [""] * (len(headers) - 1))
    ws.append([""] * len(headers))
    ws.append(headers)


def build_metrics(unit: Unit, rng: random.Random) -> dict[str, dict[str, float]]:
    revenue_budget = (380 + rng.uniform(0, 520)) * unit.scale
    revenue_factor = rng.uniform(.88, 1.14)
    if unit.index == 113:
        revenue_factor = .71
    revenue = revenue_budget * revenue_factor
    contracts = revenue * rng.uniform(.94, 1.31)
    collection = revenue * rng.uniform(.78, 1.04)
    cost = revenue * rng.uniform(.53, .67)
    sales = revenue * rng.uniform(.055, .095)
    if unit.index == 17:
        sales *= 2.15
    admin = revenue * rng.uniform(.045, .075)
    rnd = revenue * rng.uniform(.035, .068)
    profit = max(revenue - cost - sales - admin - rnd, revenue * .025)
    headcount = max(24, round(28 + unit.scale * 32 + rng.uniform(-8, 12)))
    customers = max(4, round(revenue / rng.uniform(28, 45)))
    tickets = max(80, round(headcount * rng.uniform(7.2, 11.5)))
    if unit.index == 91:
        tickets = round(tickets * 2.4)
    deliveries = max(3, round(revenue / rng.uniform(38, 62)))

    actual = {
        "营业收入": revenue,
        "回款金额": collection,
        "新签合同额": contracts,
        "营业成本": cost,
        "销售费用": sales,
        "管理费用": admin,
        "研发投入": rnd,
        "经营利润": profit,
        "在岗人数": float(headcount),
        "新增客户": float(customers),
        "服务工单": float(tickets),
        "完成交付": float(deliveries),
    }
    budget = {
        "营业收入": revenue_budget,
        "回款金额": revenue_budget * .88,
        "新签合同额": revenue_budget * 1.08,
        "营业成本": revenue_budget * .61,
        "销售费用": revenue_budget * .075,
        "管理费用": revenue_budget * .06,
        "研发投入": revenue_budget * .05,
        "经营利润": revenue_budget * .155,
        "在岗人数": float(round(headcount * 1.03)),
        "新增客户": float(round(customers * 1.08)),
        "服务工单": float(round(tickets * .94)),
        "完成交付": float(round(deliveries * 1.06)),
    }
    out: dict[str, dict[str, float]] = {}
    for key in actual:
        is_count = key in {"在岗人数", "新增客户", "服务工单", "完成交付"}
        value = actual[key]
        out[key] = {
            "budget": round(value if is_count else budget[key], 0 if is_count else 2),
            "actual": round(value, 0 if is_count else 2),
            "last": round(value * rng.uniform(.89, 1.08), 0 if is_count else 2),
            "year": round(value * rng.uniform(.82, 1.02), 0 if is_count else 2),
        }
    return out


def sheet_overview(wb: Workbook, unit: Unit, metrics: dict[str, dict[str, float]]) -> None:
    ws = wb.active
    ws.title = "经营数据汇总"
    add_header(ws, "集团分公司月度经营数据", f"{PERIOD}｜各分公司统一填报口径",
               ["指标", "单位", "本月预算", "本月实际", "上月实际", "上年同期"])
    units_map = {
        "营业收入": "万元", "回款金额": "万元", "新签合同额": "万元", "营业成本": "万元",
        "销售费用": "万元", "管理费用": "万元", "研发投入": "万元", "经营利润": "万元",
        "在岗人数": "人", "新增客户": "家", "服务工单": "单", "完成交付": "个",
    }
    for label, vals in metrics.items():
        ws.append([label, units_map[label], vals["budget"], vals["actual"], vals["last"], vals["year"]])
    style_sheet(ws, [24, 12, 18, 18, 18, 18])
    for row in ws.iter_rows(min_row=5, min_col=3, max_col=6):
        for cell in row:
            cell.number_format = "#,##0" if ws.cell(cell.row, 2).value != "万元" else "#,##0.00"


def sheet_expenses(wb: Workbook, unit: Unit, metrics: dict[str, dict[str, float]], rng: random.Random) -> None:
    ws = wb.create_sheet("费用明细")
    add_header(ws, "月度费用明细", f"{PERIOD}｜单位：万元",
               ["费用项目", "预算", "本月实际", "上月实际", "上年同期", "口径说明"])
    sales = metrics["销售费用"]["actual"]
    admin = metrics["管理费用"]["actual"]
    rnd = metrics["研发投入"]["actual"]
    rows = [
        ("市场推广费", sales * .42, "品牌、渠道及市场活动"),
        ("销售差旅费", sales * .25, "销售人员差旅"),
        ("渠道服务费", sales * .33, "渠道合作及服务"),
        ("办公及租赁费", admin * .34, "办公场地及日常支出"),
        ("行政差旅费", admin * .18, "行政管理差旅"),
        ("咨询与专业服务费", admin * .25, "法律、审计及咨询"),
        ("会议与培训费", admin * .23, "内部会议及培训"),
        ("研发人员费用", rnd * .58, "研发人员相关支出"),
        ("云资源及工具费", rnd * .24, "研发云资源及软件工具"),
        ("测试与认证费", rnd * .18, "产品测试及认证"),
    ]
    for name, actual, note in rows:
        budget = actual * rng.uniform(.88, 1.12)
        ws.append([name, round_money(budget), round_money(actual), round_money(actual*rng.uniform(.88,1.08)),
                   round_money(actual*rng.uniform(.78,1.02)), note])
    style_sheet(ws, [25, 17, 17, 17, 17, 30])
    for row in ws.iter_rows(min_row=5, min_col=2, max_col=5):
        for cell in row: cell.number_format = "#,##0.00"


def sheet_people(wb: Workbook, unit: Unit, metrics: dict[str, dict[str, float]], rng: random.Random) -> None:
    ws = wb.create_sheet("人员情况")
    add_header(ws, "月度人员情况", f"{PERIOD}｜单位：人",
               ["部门", "月初人数", "本月入职", "本月离职", "月末人数", "说明"])
    total = int(metrics["在岗人数"]["actual"])
    shares = [("销售部",.25),("交付部",.22),("客户成功部",.17),("研发部",.18),("职能部门",.12),("管理层",.06)]
    remaining = total
    for i,(name,share) in enumerate(shares):
        ending = remaining if i == len(shares)-1 else max(1, round(total*share))
        remaining -= ending
        hires = rng.randint(0, max(1, ending//14))
        departures = rng.randint(0, max(1, ending//18))
        opening = max(1, ending-hires+departures)
        note = "人员配置正常"
        ws.append([name, opening, hires, departures, ending, note])
    style_sheet(ws, [22, 16, 16, 16, 16, 28])
    for row in ws.iter_rows(min_row=5, min_col=2, max_col=5):
        for cell in row: cell.number_format = "#,##0"


def sheet_projects(wb: Workbook, unit: Unit, metrics: dict[str, dict[str, float]], rng: random.Random) -> None:
    ws = wb.create_sheet("项目交付")
    add_header(ws, "项目交付情况", f"{PERIOD}｜单位：个",
               ["项目状态", "期初数量", "本月新增", "本月完成", "期末数量", "说明"])
    delivered = int(metrics["完成交付"]["actual"])
    delayed = 2 if unit.index == 68 else rng.choice([0,0,0,1])
    rows = [
        ("实施准备", max(2,delivered//4), max(1,delivered//5), max(1,delivered//6), "等待资源与排期"),
        ("实施进行中", max(3,delivered//2), max(2,delivered//3), max(2,delivered//3), "按计划推进"),
        ("客户验收中", max(2,delivered//3), max(1,delivered//4), max(1,delivered//3), "等待客户验收"),
        ("已完成", 0, delivered, delivered, "本月完成交付"),
        ("延期项目", delayed, 0, 0, "需重点跟进" if delayed else "无延期项目"),
    ]
    for status,opening,new,done,note in rows:
        ending=max(0,opening+new-done)
        ws.append([status,opening,new,done,ending,note])
    style_sheet(ws, [22, 16, 16, 16, 16, 28])


def sheet_receivables(wb: Workbook, unit: Unit, metrics: dict[str, dict[str, float]], rng: random.Random) -> None:
    ws = wb.create_sheet("回款情况")
    add_header(ws, "应收与回款情况", f"{PERIOD}｜单位：万元",
               ["账龄区间", "期初应收", "本月新增应收", "本月回款", "期末应收", "其中逾期"])
    revenue = metrics["营业收入"]["actual"]
    collection = metrics["回款金额"]["actual"]
    weights=[("30天以内",.48,.05),("31–60天",.25,.18),("61–90天",.16,.42),("90天以上",.11,.78)]
    if unit.index == 43:
        weights=[("30天以内",.35,.06),("31–60天",.20,.20),("61–90天",.15,.48),("90天以上",.30,.92)]
    for label,weight,overdue_rate in weights:
        opening=revenue*weight*rng.uniform(.65,.9)
        new=revenue*weight
        received=min(opening+new,collection*weight*rng.uniform(.85,1.05))
        ending=max(0,opening+new-received)
        overdue=ending*overdue_rate
        ws.append([label,round_money(opening),round_money(new),round_money(received),round_money(ending),round_money(overdue)])
    style_sheet(ws, [22, 18, 18, 18, 18, 18])
    for row in ws.iter_rows(min_row=5, min_col=2, max_col=6):
        for cell in row: cell.number_format = "#,##0.00"


def sheet_anomalies(wb: Workbook, unit: Unit) -> None:
    ws = wb.create_sheet("异常与说明")
    add_header(ws, "经营异常与事项说明", f"{PERIOD}｜各单位重点事项",
               ["检查项目", "状态", "情况说明", "责任部门", "处理进度", "计划完成时间"])
    checks = [
        ("经营目标", "经营指标在合理范围内", "经营管理部"),
        ("销售费用", "费用投入符合月度计划", "市场部"),
        ("应收回款", "应收账龄结构正常", "财务部"),
        ("项目交付", "项目按计划推进", "交付部"),
        ("服务质量", "服务工单处于正常水平", "客户成功部"),
    ]
    anomaly = ANOMALIES.get(unit.index)
    for title, normal, dept in checks:
        status="正常"; description=normal; progress="无需处理"; due="—"
        if anomaly and title == ANOMALY_CHECK[unit.index]:
            status="重点关注"; description=anomaly[1]; dept=anomaly[2]; progress="整改中"; due="2026-07-15"
        ws.append([title,status,description,dept,progress,due])
    style_sheet(ws, [20, 16, 42, 20, 18, 20])
    for row in ws.iter_rows(min_row=5):
        if row[1].value == "重点关注":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFF1E8")
                cell.font = Font(name="微软雅黑", size=10, color=AMBER, bold=cell.column == 2)


def sheet_meta(wb: Workbook, unit: Unit, rng: random.Random) -> None:
    ws = wb.create_sheet("填报信息")
    add_header(ws, "月度报表填报信息", f"{PERIOD}｜来源单位信息",
               ["字段", "填报内容", "字段", "填报内容", "字段", "填报内容"])
    day=30 + (unit.index % 2)
    hour=9 + unit.index % 9
    names=["王晨","李敏","张悦","陈宇","刘洋","周宁","赵晴","孙浩","吴桐","郑洁"]
    ws.append(["单位编码",unit.code,"单位名称",unit.name,"所属区域",unit.region])
    ws.append(["报告期间",PERIOD,"报表版本","V2026.06","填报状态","已提交"])
    ws.append(["填报人",names[unit.index%len(names)],"联系电话","内部通讯录","提交时间",f"2026-06-{day} {hour:02d}:{(unit.index*7)%60:02d}"])
    ws.append(["审核状态","已审核","审核人",names[(unit.index+3)%len(names)],"数据口径","集团统一口径"])
    style_sheet(ws, [18, 28, 18, 28, 18, 28])


def build_workbook(unit: Unit) -> tuple[Workbook, dict[str, dict[str, float]]]:
    rng=random.Random(SEED+unit.index*97)
    metrics=build_metrics(unit,rng)
    wb=Workbook()
    wb.properties.title=f"{unit.name}{PERIOD}经营月报"
    wb.properties.subject="集团分公司月度经营数据统一填报"
    wb.properties.creator="表表归一演示数据生成器"
    sheet_overview(wb,unit,metrics)
    sheet_expenses(wb,unit,metrics,rng)
    sheet_people(wb,unit,metrics,rng)
    sheet_projects(wb,unit,metrics,rng)
    sheet_receivables(wb,unit,metrics,rng)
    sheet_anomalies(wb,unit)
    sheet_meta(wb,unit,rng)
    return wb,metrics


def verify(paths: list[Path]) -> dict:
    expected_sheets=["经营数据汇总","费用明细","人员情况","项目交付","回款情况","异常与说明","填报信息"]
    signatures=[]
    for path in paths:
        wb=load_workbook(path,read_only=True,data_only=True)
        assert wb.sheetnames == expected_sheets, (path.name,wb.sheetnames)
        signatures.append(tuple((ws.title,ws.max_row,ws.max_column) for ws in wb.worksheets))
    assert len(set(signatures)) == 1, "工作簿结构不一致"
    return {"file_count":len(paths),"sheet_names":expected_sheets,"structure":signatures[0]}


def main() -> None:
    for directory in (OUTPUT, RECORDING_OUTPUT):
        if directory.exists():
            shutil.rmtree(directory)
        directory.mkdir(parents=True)
    all_units=units()
    totals: dict[str, dict[str,float]]={}
    manifest=[]
    paths=[]
    for unit in all_units:
        wb,metrics=build_workbook(unit)
        path=OUTPUT/f"{unit.name}.xlsx"
        wb.save(path)
        shutil.copy2(path, RECORDING_OUTPUT / path.name)
        paths.append(path)
        for metric,vals in metrics.items():
            totals.setdefault(metric,{k:0.0 for k in vals})
            for key,value in vals.items(): totals[metric][key]+=float(value)
        manifest.append({
            "index":unit.index,"code":unit.code,"region":unit.region,"unit":unit.name,
            "file":path.name,"anomaly":ANOMALIES.get(unit.index),
        })
    for vals in totals.values():
        for key in vals: vals[key]=round(vals[key],2)
    report=verify(paths)
    (ROOT/"manifest.json").write_text(json.dumps({"scenario":"集团总部每月收集120家分公司同构经营月报", "period":PERIOD,
        "verification":report,"units":manifest,"expected_group_totals":totals},ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps({"output":str(OUTPUT),"recording_output":str(RECORDING_OUTPUT),
        "verification":report,"anomalies":[manifest[i-1] for i in ANOMALIES]},ensure_ascii=False,indent=2))


if __name__ == "__main__":
    main()
